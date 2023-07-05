from support import SupportSubprocess
from .setup import *
from .utils import *
from .models import *

name = 'rcpt'

@app.route('/download_sample', methods=['GET'])
def download_sample():
    from flask import send_file
    fpath = '/data/csmail/files/userlist_sample.xlsx'
    excel = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return send_file(fpath, mimetype=excel, download_name='userinfo_sample.xlsx', as_attachment=True)

class ModuleRcpt(PluginModuleBase):

    def __init__(self, P):
        super(ModuleRcpt, self).__init__(P, name='rcpt', first_menu='list')
        self.web_list_model = ModelRcptItem
        self.set_page_list([PageRcptList])
        self.curr_rcpt_list = []

    def process_command(self, command, arg1, arg2, arg3, req):
        ret = {}
        data = None
        logger.info(f'[{self.name}] process_command: {command}, {arg1}, {req}')
        if command == 'include_rcpt' or command == 'exclude_rcpt':
            rcpt = ModelRcptItem.get_by_id(arg1)
            rcptlist = ModelRcptListItem.get_by_id(rcpt.id)
            rcpt.excluded = False if command == 'include_rcpt' else True
            if rcpt.excluded:
                rcptlist.except_count += 1
                rcptlist.target_count -= 1
            else:
                rcptlist.except_count -= 1
                rcptlist.target_count += 1
            rcpt.save()
            rcptlist.save()
            ret = {'ret':'success', 'data':'완료하였습니다'}
        elif command == 'register_rcpt':
            _ = P.logic.arg_to_dict(arg1)
            rcpt = ModelRcptItem(int(_['rcpt_option']), _['name'], _['dept'], _['email'], _['emp_code'])
            rcpt.save()
            rcptlist = ModelRcptListItem.get_by_id(int(_['rcpt_option']))
            rcptlist.total_count += 1
            rcptlist.target_count += 1
            rcptlist.save()
            ret = {'ret':'success', 'data':f'대상자({rcpt.name}) 등록 완료'}
        elif command == 'modify_rcpt':
            _ = P.logic.arg_to_dict(arg1)
            rcpt = ModelRcptItem.get_by_id(int(_['m_rcpt_id']))
            rcpt.name = _['m_name']
            rcpt.dept = _['m_dept']
            rcpt.list_id = _['m_rcpt_option']
            rcpt.emp_code = _['m_emp_code']
            rcpt.email = _['m_email']
            rcpt.save()
            ret = {'ret':'success', 'data':f'대상자({rcpt.name}) 수정 완료'}
        elif command == 'remove_rcpt':
            try:
                rcpt = ModelRcptItem.get_by_id(int(arg1))
                if rcpt.list_id != -1:
                    rcptlist = ModelRcptListItem.get_by_id(rcpt.list_id)
                    rcptlist.total_count -= 1
                    if rcpt.excluded: rcptlist.except_count -= 1
                    else: rcptlist.target_count -= 1
                    rcptlist.save()
                ModelRcptItem.delete_by_id(arg1)
                ret = {'ret':'success', 'data':f'삭제 완료(ID:{arg1})'}
            except Exception as e:
                logger.error('Exception:%s', e)
                logger.error(traceback.format_exc())
                ret = {'ret':'error', 'data':f'삭제 실패(ID:{arg1})'}
        return jsonify(ret)

    def process_menu(self, page, req):
        arg = P.ModelSetting.to_dict()
        rcptlist = ModelRcptListItem.get_all_rcptlist()
        deptlist = list(x.dept for x in ModelRcptItem.get_all_depts())
        logger.info(f'[depts] {deptlist}')
        logger.info(f'[{self.name}] req({req}, {req.args})')
        if 'rcptlist_id' in req.form: arg['start_rcptlist_id'] = req.form['rcptlist_id']
        return render_template(f'{self.P.package_name}_{self.name}_{page}.html', arg=arg, rcptlist=rcptlist, deptlist=deptlist)

class PageRcptList(PluginPageBase):
    def __init__(self, P, parent):
        super(PageRcptList, self).__init__(P, parent, name='list')
        self.curr_rcpt_list = []

    def process_ajax(self, sub, req):
        ret = {'ret':'success'}
        logger.info(f'[rcptlist] ajax({sub}), {req.form}, {req.files}')
        if sub == 'upload_user_list':
            from openpyxl import load_workbook
            from io import BytesIO
            except_depts = list(filter(None, list(x.strip() for x in req.form.get('except_dept_names').split(','))))
            upfile = req.files.get('upload_file')
            wb = load_workbook(filename = BytesIO(upfile.read()))
            sheet = wb.active
            user_list = []
            for row in sheet.iter_rows(min_row=2):
                user_data = {'emp_code': row[0].value, 'name':row[1].value, 'dept_name':row[2].value, 'email':row[3].value, 'excluded':False}
                if row[2].value in except_depts: user_data['excluded'] = True
                user_list.append(user_data)
            logger.debug(f'[rcpt_list] user_list: {user_list}')
            self.curr_rcpt_list = user_list
            ret['data'] = user_list
            ret['ret'] = 'success'
        return jsonify(ret)

    def process_command(self, command, arg1, arg2, arg3, req):
        logger.info(f'[rcptlist] command({command}), {arg1}, {arg2}, {arg3}')
        ret = {}
        if command == 'get_user_list':
            except_depts = []
            if arg1 != '': except_depts = list(filter(None, list(x.strip() for x in arg1.split(','))))
            logger.debug(f'[rcptlist] except_depts: {except_depts}')
            ret = ScUtils.get_user_list(except_depts)
            if ret['ret'] == 'success':
                self.curr_rcpt_list = ret['data']
        elif command == 'register_rcptlist':
            ret = self.register_rcpt_list(P.logic.arg_to_dict(arg1))
        elif command == 'refresh_rcptlist':
            ret = {'ret':'success', 'data':'건수 갱신 성공'}
            rcptlist_id = int(arg1)
            rcptlist = ModelRcptListItem.get_by_id(rcptlist_id)
            rcpts = ModelRcptItem.get_by_list_id(rcptlist_id)
            rcptlist.total_count = len(rcpts)
            rcptlist.target_count = 0
            rcptlist.except_count = 0
            for rcpt in rcpts:
                if rcpt.excluded: rcptlist.except_count += 1
                else: rcptlist.target_count += 1
            rcptlist.save()
        elif command == 'remove_rcptlist':
            try:
                ModelRcptItem.delete_by_list_id(int(arg1))
                ModelRcptListItem.delete_by_id(int(arg1))
                ret = {'ret':'success', 'data':f'삭제 완료(ID:{arg1})'}
            except Exception as e:
                logger.error('Exception:%s', e)
                logger.error(traceback.format_exc())
                ret = {'ret':'error', 'data':f'{str(e)}'}
        elif command == 'web_list':
            ret = ModelRcptListItem.web_list(P.logic.arg_to_dict(arg1))
        return jsonify(ret)

    def register_rcpt_list(self, req):
        if len(self.curr_rcpt_list) == 0 and req['empty_list'] == False:
            return {'ret':'error', 'data':'먼저파일을 선택하거나, 사용자정보가져오기 후 실행해주세요'}

        except_depts = list(filter(None, list(x.strip() for x in req['except_dept_names'].split(','))))
        name = req['name']
        rcptlist = ModelRcptListItem(name)
        rcptlist.total_count = len(self.curr_rcpt_list)
        rcptlist.target_count = rcptlist.total_count
        rcptlist.except_count = 0
        rcptlist.save()

        for _ in self.curr_rcpt_list:
            rcpt = ModelRcptItem(rcptlist.id, _['name'], _['dept_name'], _['email'], _['emp_code'])
            if _['dept_name'] in except_depts or _['excluded']:
                rcptlist.except_count += 1
                rcptlist.target_count -= 1
                rcpt.excluded = True
            rcpt.save()

        self.curr_rcpt_list = []
        rcptlist.save()
        return {'ret':'success', 'data':f'수신자그룹({name}) 등록완료: 전체({rcptlist.total_count}),등록({rcptlist.target_count}),제외({rcptlist.except_count})'}
